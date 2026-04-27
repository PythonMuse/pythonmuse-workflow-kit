"""
export_excel.py
Exports the full joined dataset with live Excel formulas for gross profit and margin %.
Output: /outputs/gross_margin_analysis.xlsx
"""

import csv
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

RAW_DIR   = os.path.join(os.path.dirname(__file__), '..', 'data', 'raw')
OUT_DIR   = os.path.join(os.path.dirname(__file__), '..', 'outputs')
REV_FILE  = os.path.join(RAW_DIR, 'pythonmuse_orders_revenue.csv')
COST_FILE = os.path.join(RAW_DIR, 'pythonmuse_orders_costs.csv')
OUT_FILE  = os.path.join(OUT_DIR, 'gross_margin_analysis.xlsx')

HEADERS = [
    # order info
    'order_id', 'order_date', 'customer', 'region', 'product', 'salesperson',
    # revenue
    'quantity', 'unit_price', 'total_revenue',
    # costs
    'vendor', 'material_cost', 'labor_employee', 'labor_hours', 'labor_rate', 'labor_cost', 'total_cogs',
    # calculated
    'gross_profit', 'margin_pct',
]

HEADER_LABELS = [
    'Order ID', 'Order Date', 'Customer', 'Region', 'Product', 'Salesperson',
    'Qty', 'Unit Price', 'Total Revenue',
    'Vendor', 'Material Cost', 'Labor Employee', 'Labor Hours', 'Labor Rate', 'Labor Cost', 'Total COGS',
    'Gross Profit', 'Margin %',
]

# Column indices (1-based)
COL = {h: i + 1 for i, h in enumerate(HEADERS)}

CURRENCY_COLS = {'unit_price', 'total_revenue', 'material_cost', 'labor_cost', 'total_cogs', 'gross_profit', 'labor_rate'}
INT_COLS      = {'quantity', 'labor_hours'}

def read_csv(path):
    with open(path, newline='', encoding='utf-8') as f:
        return list(csv.DictReader(f))

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def main():
    rev   = read_csv(REV_FILE)
    costs = read_csv(COST_FILE)
    costs_by_id = {r['order_id']: r for r in costs}

    # join and sort by margin (revenue - cogs) / revenue ascending
    rows = []
    for r in rev:
        c = costs_by_id[r['order_id']]
        margin = (float(r['total_revenue']) - float(c['total_cogs'])) / float(r['total_revenue'])
        rows.append((margin, r, c))
    rows.sort(key=lambda x: x[0])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Gross Margin Analysis'

    # ── Metadata block ────────────────────────────────────────────────────────
    meta_font = Font(name='Calibri', size=9, color='666666')
    ws['A1'] = 'PythonMuse Gross Margin Analysis'
    ws['A1'].font = Font(name='Calibri', size=12, bold=True)
    ws['A2'] = 'Period: Jan 2024 – Nov 2025'
    ws['A2'].font = meta_font
    ws['A3'] = f'Generated: {datetime.today().strftime("%Y-%m-%d")}'
    ws['A3'].font = meta_font
    ws['A4'] = 'Gross Profit = Total Revenue − Total COGS'
    ws['A4'].font = meta_font
    ws['A5'] = 'Margin % = Gross Profit / Total Revenue'
    ws['A5'].font = meta_font

    HEADER_ROW = 7

    # ── Header row ────────────────────────────────────────────────────────────
    header_fill   = PatternFill('solid', fgColor='1F4E79')
    calc_fill     = PatternFill('solid', fgColor='2E75B6')
    header_font   = Font(name='Calibri', size=10, bold=True, color='FFFFFF')

    for col_idx, label in enumerate(HEADER_LABELS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=label)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border()
        # highlight calculated columns differently
        if label in ('Gross Profit', 'Margin %'):
            cell.fill = calc_fill
        else:
            cell.fill = header_fill

    # ── Data rows ─────────────────────────────────────────────────────────────
    fmt_currency = '#,##0.00'
    fmt_pct      = '0.00%'
    fmt_date     = 'YYYY-MM-DD'

    alt_fill  = PatternFill('solid', fgColor='EBF3FB')
    even_fill = PatternFill('solid', fgColor='FFFFFF')

    for row_offset, (_, r, c) in enumerate(rows):
        row_num  = HEADER_ROW + 1 + row_offset
        row_fill = alt_fill if row_offset % 2 == 1 else even_fill

        rev_col  = get_column_letter(COL['total_revenue'])
        cogs_col = get_column_letter(COL['total_cogs'])
        gp_col   = get_column_letter(COL['gross_profit'])

        values = {
            'order_id':       r['order_id'],
            'order_date':     datetime.strptime(r['order_date'], '%Y-%m-%d'),
            'customer':       r['customer'],
            'region':         r['region'],
            'product':        r['product'],
            'salesperson':    r['salesperson'],
            'quantity':       int(r['quantity']),
            'unit_price':     float(r['unit_price']),
            'total_revenue':  float(r['total_revenue']),
            'vendor':         c['vendor'],
            'material_cost':  float(c['material_cost']),
            'labor_employee': c['labor_employee'],
            'labor_hours':    int(c['labor_hours']),
            'labor_rate':     float(c['labor_rate']),
            'labor_cost':     float(c['labor_cost']),
            'total_cogs':     float(c['total_cogs']),
            # live formulas
            'gross_profit':   f'={rev_col}{row_num}-{cogs_col}{row_num}',
            'margin_pct':     f'={gp_col}{row_num}/{rev_col}{row_num}',
        }

        for col_name, value in values.items():
            col_idx = COL[col_name]
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = Font(name='Calibri', size=10)
            cell.fill = row_fill
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal='center' if col_name in
                ('order_id', 'order_date', 'quantity', 'labor_hours', 'region', 'product') else 'left')

            if col_name == 'order_date':
                cell.number_format = fmt_date
            elif col_name in CURRENCY_COLS:
                cell.number_format = fmt_currency
            elif col_name == 'margin_pct':
                cell.number_format = fmt_pct

    # ── Totals row ────────────────────────────────────────────────────────────
    DATA_START = HEADER_ROW + 1
    DATA_END   = HEADER_ROW + len(rows)
    totals_row = DATA_END + 1

    totals_fill = PatternFill('solid', fgColor='D6E4F0')
    totals_font = Font(name='Calibri', size=10, bold=True)

    sum_cols    = {'total_revenue', 'material_cost', 'labor_cost', 'total_cogs', 'gross_profit'}
    label_col   = COL['order_id']

    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=totals_row, column=col_idx)
        cell.fill = totals_fill
        cell.font = totals_font
        cell.border = thin_border()

    ws.cell(row=totals_row, column=label_col, value='TOTAL').alignment = Alignment(horizontal='center')

    for col_name in sum_cols:
        col_idx  = COL[col_name]
        col_lett = get_column_letter(col_idx)
        cell = ws.cell(row=totals_row, column=col_idx,
                       value=f'=SUM({col_lett}{DATA_START}:{col_lett}{DATA_END})')
        cell.number_format = fmt_currency
        cell.alignment = Alignment(horizontal='right')

    # overall margin formula in totals row
    gp_lett  = get_column_letter(COL['gross_profit'])
    rev_lett = get_column_letter(COL['total_revenue'])
    mgn_cell = ws.cell(row=totals_row, column=COL['margin_pct'],
                       value=f'={gp_lett}{totals_row}/{rev_lett}{totals_row}')
    mgn_cell.number_format = fmt_pct

    # ── Conditional formatting on margin % ───────────────────────────────────
    mgn_col   = get_column_letter(COL['margin_pct'])
    mgn_range = f'{mgn_col}{DATA_START}:{mgn_col}{totals_row}'
    ws.conditional_formatting.add(
        mgn_range,
        ColorScaleRule(
            start_type='min',  start_color='F4CCCC',   # red
            mid_type='percentile', mid_value=50, mid_color='FFE599',  # yellow
            end_type='max',    end_color='B7E1CD',     # green
        )
    )

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = {
        'order_id': 9, 'order_date': 12, 'customer': 18, 'region': 11,
        'product': 13, 'salesperson': 14, 'quantity': 7, 'unit_price': 10,
        'total_revenue': 14, 'vendor': 22, 'material_cost': 13,
        'labor_employee': 15, 'labor_hours': 11, 'labor_rate': 10,
        'labor_cost': 11, 'total_cogs': 11, 'gross_profit': 13, 'margin_pct': 10,
    }
    for col_name, width in widths.items():
        ws.column_dimensions[get_column_letter(COL[col_name])].width = width

    ws.row_dimensions[HEADER_ROW].height = 30

    # ── Freeze panes and zoom ─────────────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=1)
    ws.sheet_view.zoomScale = 90

    wb.save(OUT_FILE)
    print(f'Saved: {OUT_FILE}')

if __name__ == '__main__':
    main()
